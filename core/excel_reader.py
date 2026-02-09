"""
Excel 读取模块 - 处理 Excel 文件导入
"""

from openpyxl import load_workbook
from typing import List, Optional
import os


class ExcelReader:
    """Excel 文件读取器"""
    
    def __init__(self):
        self.workbook = None
        self.sheet = None
        self.file_path = None
    
    def load_file(self, file_path: str) -> tuple[bool, str]:
        """
        加载 Excel 文件
        
        Args:
            file_path: Excel 文件路径
            
        Returns:
            (是否成功, 消息)
        """
        try:
            if not os.path.exists(file_path):
                return False, "文件不存在"
            
            self.workbook = load_workbook(file_path, read_only=True, data_only=True)
            self.sheet = self.workbook.active
            self.file_path = file_path
            
            return True, f"成功加载文件，共 {self.sheet.max_row - 1} 行数据"
            
        except Exception as e:
            return False, f"加载失败: {str(e)}"
    
    def get_columns(self) -> List[str]:
        """
        获取第一行的列名（表头）
        
        Returns:
            列名列表
        """
        if not self.sheet:
            return []
        
        columns = []
        for cell in self.sheet[1]:
            value = cell.value
            if value:
                columns.append(str(value))
        
        return columns
    
    def get_column_index(self, column_name: str) -> Optional[int]:
        """
        根据列名获取列索引（1-based）
        
        Args:
            column_name: 列名
            
        Returns:
            列索引，未找到返回 None
        """
        columns = self.get_columns()
        try:
            return columns.index(column_name) + 1
        except ValueError:
            return None
    
    def get_titles(self, column_name: str) -> List[str]:
        """
        获取指定列的所有标题（跳过第一行表头）
        
        Args:
            column_name: 列名
            
        Returns:
            标题列表
        """
        if not self.sheet:
            return []
        
        col_index = self.get_column_index(column_name)
        if col_index is None:
            return []
        
        titles = []
        for row_num in range(2, self.sheet.max_row + 1):
            cell_value = self.sheet.cell(row=row_num, column=col_index).value
            if cell_value:
                titles.append(str(cell_value).strip())
        
        return titles
    
    def get_row_count(self) -> int:
        """获取数据行数（不含表头）"""
        if not self.sheet:
            return 0
        return max(0, self.sheet.max_row - 1)
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.sheet = None
